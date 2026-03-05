#!/usr/bin/env python3
import json
from pathlib import Path
from warnings import filters

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from fill_expected_topk import parse_filters


BASE = Path(__file__).resolve().parent.parent
IN_TESTS = BASE / "out" / "analysis" / "random_testcases_with_expected.csv"
LOG_PATH = BASE / "out" / "vigade_log.csv"
OUT_XLSX = BASE / "out" / "analysis" / "testjuhtumid.xlsx"


def safe_json_loads(x: str):
    try:
        return json.loads(x) if isinstance(x, str) else {}
    except Exception:
        return {}


def split_codes(s: str):
    if not isinstance(s, str) or not s.strip():
        return []
    return [c.strip() for c in s.split(",") if c.strip()]


def make_xlsx(df: pd.DataFrame, out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Testjuhtumid"

    headers = list(df.columns)
    ws.append(headers)
    for _, row in df.iterrows():
        ws.append([row.get(h, "") for h in headers])

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"

    widths = {
        "A": 6,   # ID
        "B": 55,  # Päring
        "C": 45,  # Filtrid
        "D": 70,  # Expected
        "E": 18,  # PASS/FAIL
        "F": 40,  # Märkus
        "G": 18,  # Logi aeg
        "H": 12,  # Logi tulemus
        "I": 70,  # Logi top_codes
        "J": 18,  # Match?
    }
    for i in range(1, len(headers) + 1):
        col_letter = get_column_letter(i)
        if col_letter in widths:
            ws.column_dimensions[col_letter].width = widths[col_letter]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            if cell.column_letter in {"B", "C", "D", "F", "I"}:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    table = Table(displayName="TestJuhtumid", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main():
    if not IN_TESTS.exists():
        raise SystemExit(f"Puudub: {IN_TESTS}. Käivita enne generate_random_testcases.py")

    if not LOG_PATH.exists():
        raise SystemExit(f"Puudub: {LOG_PATH}. Käivita päringud rakenduses, et log tekiks.")

    tests = pd.read_csv(IN_TESTS).fillna("")
    log = pd.read_csv(LOG_PATH).fillna("")

    required_log = {"Aeg", "Päring", "Filtrid", "Tulemus", "DetailidJSON"}
    missing = required_log - set(log.columns)
    if missing:
        raise SystemExit(f"vigade_log.csv puuduvad veerud: {sorted(missing)}")

    details = log["DetailidJSON"].apply(safe_json_loads)
    log["top_codes"] = details.apply(lambda d: d.get("top_codes", []))
    log["top_codes_str"] = log["top_codes"].apply(lambda xs: ", ".join(xs) if isinstance(xs, list) else "")

    # Match newest by (Päring, Filtrid)
    def parse_filters(filters_str: str):
        out = {}
        for part in str(filters_str).split(","):
            if "=" in part:
                k, v = part.split("=", 1)
                out[k.strip()] = v.strip()
        return out

    #   Match newest by (Päring + core filters). Ignore extra keys present only in tests.
    CORE_KEYS = {"credits", "semester", "language", "level"}

    def find_match(query: str, filters: str):
        tf = parse_filters(filters)
        # candidates with same query
        m = log[log["Päring"].astype(str).str.strip() == query.strip()]
        if len(m) == 0:
            return None

        # keep only those where core keys match (if present in log)
        def ok(row):
            lf = parse_filters(row.get("Filtrid", ""))
            for k in CORE_KEYS:
                if k in lf and k in tf and str(lf[k]).strip() != str(tf[k]).strip():
                    return False
            return True

        m = m[m.apply(ok, axis=1)]
        if len(m) == 0:
            return None

        return m.sort_values("Aeg", ascending=False).iloc[0]

    out = tests.copy()
    log_time, log_result, log_top = [], [], []
    hit_counts, exp_counts, log_counts = [], [], []

    for _, r in out.iterrows():
        m = find_match(str(r["Päring"]), str(r["Filtrid"]))
        if m is None:
            log_time.append("")
            log_result.append("")
            log_top.append("")
            hit_counts.append(0)
            exp_counts.append(len(split_codes(r.get("Expected unique_ID (top_codes)", ""))))
            log_counts.append(0)
            continue

        log_time.append(m.get("Aeg", ""))
        log_result.append(m.get("Tulemus", ""))
        log_top_codes = m.get("top_codes", [])
        log_top.append(m.get("top_codes_str", ""))

        exp = split_codes(r.get("Expected unique_ID (top_codes)", ""))
        overlap = sorted(set(exp).intersection(set(log_top_codes))) if isinstance(log_top_codes, list) else []
        hit_counts.append(len(overlap))
        exp_counts.append(len(exp))
        log_counts.append(len(log_top_codes) if isinstance(log_top_codes, list) else 0)

    out["Logi aeg"] = log_time
    out["Logi tulemus"] = log_result
    out["Logi top_codes"] = log_top
    out["Expected vs Logi"] = [f"{h}/{e}" for h, e in zip(hit_counts, exp_counts)]
    out["Expected"] = exp_counts
    out["Logi"] = log_counts

    # PASS/FAIL:
    # - FAIL if no matching log row
    # - FAIL if log_result is BAD
    # - otherwise PASS if HIT else FAIL
    def passfail(row):
        if row["Logi"] == 0:
            return "FAIL"
        if str(row.get("Logi tulemus", "")).upper() == "BAD":
            return "FAIL"
        hit_str = str(row.get("Expected vs Logi", "0/0"))
        try:
            hits = int(hit_str.split("/", 1)[0])
        except Exception:
            hits = 0
        return "PASS" if hits > 0 else "FAIL"

    out["Tulemus (PASS/FAIL)"] = out.apply(passfail, axis=1)

    # Notes for failures
    out.loc[out["Logi"] == 0, "Märkus"] = (
        out.loc[out["Logi"] == 0, "Märkus"]
        .replace("", "Ei leidnud vastavat rida vigade_log.csv-st (päring+filtrid peavad täpselt klappima).")
    )
    out.loc[(out["Logi"] > 0) & (out["Expected vs Logi"] == 0), "Märkus"] = (
        out.loc[(out["Logi"] > 0) & (out["Expected vs Logi"] == 0), "Märkus"]
        .replace("", "Logi top_codes ei sisaldanud Expected koodidest ühtegi (HIT puudub).")
    )

    # Reorder columns: ID, Expected vs Logi, Expected, Logi, then the rest
    first_cols = ["ID", "Expected vs Logi", "Expected", "Logi"]
    remaining_cols = [c for c in out.columns if c not in first_cols]
    out = out[first_cols + remaining_cols]

    make_xlsx(out, OUT_XLSX)
    print(f"Valmis: {OUT_XLSX}")


if __name__ == "__main__":
    main()
